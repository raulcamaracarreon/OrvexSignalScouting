from __future__ import annotations

import csv
import os
import unicodedata
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from urllib.parse import urlencode

import requests
from dotenv import load_dotenv
from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent
ENV_PATH = PROJECT_ROOT / ".env"

DENUE_BASE_URL = (
    "https://www.inegi.org.mx/app/api/denue/v1/consulta"
)

DENUE_BATCH_SIZE = 500

COVERAGE_OPTIONS = [
    {
        "value": "500",
        "label": "Rápida — hasta 500 establecimientos",
    },
    {
        "value": "1000",
        "label": "Equilibrada — hasta 1,000 establecimientos",
    },
    {
        "value": "2000",
        "label": "Amplia — hasta 2,000 establecimientos",
    },
    {
        "value": "5000",
        "label": "Exhaustiva — hasta 5,000 establecimientos",
    },
]

VALID_COVERAGE_LIMITS = {
    int(option["value"])
    for option in COVERAGE_OPTIONS
}

ENTITY_OPTIONS = [
    {
        "code": "09",
        "name": "Ciudad de México",
    },
]

MUNICIPALITY_OPTIONS = [
    {
        "code": "0",
        "name": "Todas las alcaldías",
    },
    {
        "code": "002",
        "name": "Azcapotzalco",
    },
    {
        "code": "003",
        "name": "Coyoacán",
    },
    {
        "code": "004",
        "name": "Cuajimalpa de Morelos",
    },
    {
        "code": "005",
        "name": "Gustavo A. Madero",
    },
    {
        "code": "006",
        "name": "Iztacalco",
    },
    {
        "code": "007",
        "name": "Iztapalapa",
    },
    {
        "code": "008",
        "name": "La Magdalena Contreras",
    },
    {
        "code": "009",
        "name": "Milpa Alta",
    },
    {
        "code": "010",
        "name": "Álvaro Obregón",
    },
    {
        "code": "011",
        "name": "Tláhuac",
    },
    {
        "code": "012",
        "name": "Tlalpan",
    },
    {
        "code": "013",
        "name": "Xochimilco",
    },
    {
        "code": "014",
        "name": "Benito Juárez",
    },
    {
        "code": "015",
        "name": "Cuauhtémoc",
    },
    {
        "code": "016",
        "name": "Miguel Hidalgo",
    },
    {
        "code": "017",
        "name": "Venustiano Carranza",
    },
]

STRATUM_OPTIONS = [
    {
        "code": "0",
        "name": "Todos los tamaños",
    },
    {
        "code": "1",
        "name": "0 a 5 personas",
    },
    {
        "code": "2",
        "name": "6 a 10 personas",
    },
    {
        "code": "3",
        "name": "11 a 30 personas",
    },
    {
        "code": "4",
        "name": "31 a 50 personas",
    },
    {
        "code": "5",
        "name": "51 a 100 personas",
    },
    {
        "code": "6",
        "name": "101 a 250 personas",
    },
    {
        "code": "7",
        "name": "251 y más personas",
    },
]

NORMALIZED_FIELDS = (
    "Nombre",
    "Razon_social",
    "Clase_actividad",
    "Estrato",
    "Tipo_vialidad",
    "Calle",
    "Num_Exterior",
    "Num_Interior",
    "Colonia",
    "CP",
    "Ubicacion",
    "Telefono",
    "Correo_e",
    "Sitio_internet",
    "Longitud",
    "Latitud",
)

EXPORT_COLUMNS = (
    (
        "Negocio",
        "Nombre",
    ),
    (
        "Razón social",
        "Razon_social",
    ),
    (
        "Actividad",
        "Clase_actividad",
    ),
    (
        "Estrato",
        "Estrato",
    ),
    (
        "Teléfono",
        "Telefono",
    ),
    (
        "Correo",
        "Correo_e",
    ),
    (
        "Sitio web",
        "Sitio_internet",
    ),
    (
        "Calle",
        "Calle",
    ),
    (
        "Número exterior",
        "Num_Exterior",
    ),
    (
        "Colonia",
        "Colonia",
    ),
    (
        "Código postal",
        "CP",
    ),
    (
        "Ubicación",
        "Ubicacion",
    ),
    (
        "Google Maps",
        "Google_Maps_URL",
    ),
)

load_dotenv(ENV_PATH)

app = Flask(__name__)


def get_default_form() -> dict[str, object]:
    return {
        "entity": "09",
        "municipality": "007",
        "activity_class": "468213",
        "stratum": "0",
        "maximum_records": "1000",
        "colony_filter": "",
        "only_with_phone": True,
        "only_without_website": True,
    }


def get_valid_codes(
    options: list[dict[str, str]],
) -> set[str]:
    return {
        option["code"]
        for option in options
    }


def get_option_name(
    options: list[dict[str, str]],
    selected_code: str,
) -> str:
    for option in options:
        if option["code"] == selected_code:
            return option["name"]

    return selected_code


def query_denue_range(
    session: requests.Session,
    entity: str,
    municipality: str,
    activity_class: str,
    stratum: str,
    initial_record: int,
    final_record: int,
) -> list[dict[str, object]]:
    token = os.getenv("DENUE_TOKEN", "").strip()

    if not token:
        raise RuntimeError(
            "No se encontró DENUE_TOKEN en el archivo .env."
        )

    request_url = (
        f"{DENUE_BASE_URL}/BuscarAreaActEstr/"
        f"{entity}/"
        f"{municipality}/"
        "0/"
        "0/"
        "0/"
        "0/"
        "0/"
        "0/"
        f"{activity_class}/"
        "0/"
        f"{initial_record}/"
        f"{final_record}/"
        "0/"
        f"{stratum}/"
        f"{token}"
    )

    response = session.get(
        request_url,
        timeout=30,
        headers={
            "Accept": "application/json",
            "User-Agent": "OrvexSignalScouting/0.3",
        },
    )

    response.raise_for_status()

    data = response.json()

    if not isinstance(data, list):
        raise RuntimeError(
            "La API del DENUE devolvió una estructura inesperada."
        )

    return [
        record
        for record in data
        if isinstance(record, dict)
    ]


def query_denue_coverage(
    entity: str,
    municipality: str,
    activity_class: str,
    stratum: str,
    coverage_limit: int,
) -> tuple[list[dict[str, object]], int, bool]:
    accumulated_records: list[dict[str, object]] = []

    batch_count = 0
    initial_record = 1

    with requests.Session() as session:
        while initial_record <= coverage_limit:
            final_record = min(
                initial_record + DENUE_BATCH_SIZE - 1,
                coverage_limit,
            )

            requested_batch_size = (
                final_record - initial_record + 1
            )

            batch_records = query_denue_range(
                session=session,
                entity=entity,
                municipality=municipality,
                activity_class=activity_class,
                stratum=stratum,
                initial_record=initial_record,
                final_record=final_record,
            )

            batch_count += 1

            accumulated_records.extend(
                batch_records[:requested_batch_size]
            )

            if len(batch_records) < requested_batch_size:
                break

            initial_record = final_record + 1

    coverage_limit_reached = (
        len(accumulated_records) >= coverage_limit
    )

    return (
        accumulated_records,
        batch_count,
        coverage_limit_reached,
    )


def normalize_text(value: object) -> str:
    if value is None:
        return ""

    return " ".join(str(value).split())


def normalize_search_text(value: object) -> str:
    normalized_value = unicodedata.normalize(
        "NFD",
        normalize_text(value),
    )

    without_accents = "".join(
        character
        for character in normalized_value
        if unicodedata.category(character) != "Mn"
    )

    return without_accents.casefold()


def build_google_maps_url(
    record: dict[str, object],
) -> str:
    query_parts = [
        normalize_text(record.get("Nombre")),
        normalize_text(record.get("Calle")),
        normalize_text(record.get("Num_Exterior")),
        normalize_text(record.get("Colonia")),
        normalize_text(record.get("Ubicacion")),
    ]

    query = ", ".join(
        part
        for part in query_parts
        if part
    )

    if not query:
        latitude = normalize_text(
            record.get("Latitud")
        )

        longitude = normalize_text(
            record.get("Longitud")
        )

        if latitude and longitude:
            query = f"{latitude},{longitude}"

    if not query:
        return ""

    parameters = urlencode(
        {
            "api": "1",
            "query": query,
        }
    )

    return (
        "https://www.google.com/maps/search/"
        f"?{parameters}"
    )

def normalize_record(
    record: dict[str, object],
) -> dict[str, object]:
    normalized_record = dict(record)

    for field_name in NORMALIZED_FIELDS:
        normalized_record[field_name] = normalize_text(
            record.get(field_name)
        )

    normalized_record["Google_Maps_URL"] = (
        build_google_maps_url(
            normalized_record
        )
    )

    return normalized_record


def filter_results(
    records: list[dict[str, object]],
    only_with_phone: bool,
    only_without_website: bool,
    colony_filter: str,
) -> list[dict[str, object]]:
    filtered_records: list[dict[str, object]] = []

    normalized_colony_filter = normalize_search_text(
        colony_filter
    )

    for record in records:
        normalized_record = normalize_record(record)

        phone = normalize_text(
            normalized_record.get("Telefono")
        )

        website = normalize_text(
            normalized_record.get("Sitio_internet")
        )

        colony = normalize_search_text(
            normalized_record.get("Colonia")
        )

        if only_with_phone and not phone:
            continue

        if only_without_website and website:
            continue

        if (
            normalized_colony_filter
            and normalized_colony_filter not in colony
        ):
            continue

        filtered_records.append(normalized_record)

    return sorted(
        filtered_records,
        key=lambda item: normalize_search_text(
            item.get("Nombre")
        ),
    )

def get_export_value(
    record: dict[str, object],
    field_name: str,
) -> str:
    value = normalize_text(
        record.get(field_name)
    )

    if value.startswith(
        (
            "=",
            "+",
            "-",
            "@",
        )
    ):
        return f"'{value}"

    return value


def build_export_filename(
    extension: str,
) -> str:
    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    return (
        "orvexsignal_prospectos_"
        f"{timestamp}.{extension}"
    )


def export_results_csv(
    results: list[dict[str, object]],
):
    text_stream = StringIO(
        newline="",
    )

    writer = csv.writer(
        text_stream,
    )

    writer.writerow(
        [
            column_title
            for column_title, _ in EXPORT_COLUMNS
        ]
    )

    for record in results:
        writer.writerow(
            [
                get_export_value(
                    record,
                    field_name,
                )
                for _, field_name in EXPORT_COLUMNS
            ]
        )

    csv_content = (
        "\ufeff"
        + text_stream.getvalue()
    ).encode("utf-8")

    byte_stream = BytesIO(
        csv_content
    )

    byte_stream.seek(0)

    return send_file(
        byte_stream,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=build_export_filename(
            "csv"
        ),
    )


def export_results_excel(
    results: list[dict[str, object]],
):
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Prospectos"
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    column_titles = [
        column_title
        for column_title, _ in EXPORT_COLUMNS
    ]

    worksheet.append(
        column_titles
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="312E81",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    worksheet.row_dimensions[1].height = 30

    for record in results:
        worksheet.append(
            [
                get_export_value(
                    record,
                    field_name,
                )
                for _, field_name in EXPORT_COLUMNS
            ]
        )

    content_alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )

    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = content_alignment

    column_widths = {
        "Negocio": 30,
        "Razón social": 30,
        "Actividad": 48,
        "Estrato": 20,
        "Teléfono": 18,
        "Correo": 30,
        "Sitio web": 32,
        "Calle": 28,
        "Número exterior": 16,
        "Colonia": 28,
        "Código postal": 15,
        "Ubicación": 38,
        "Google Maps": 42,
    }

    for column_number, column_title in enumerate(
        column_titles,
        start=1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        worksheet.column_dimensions[
            column_letter
        ].width = column_widths.get(
            column_title,
            20,
        )

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    website_column = (
        column_titles.index("Sitio web")
        + 1
    )

    maps_column = (
        column_titles.index("Google Maps")
        + 1
    )

    for row_number, record in enumerate(
        results,
        start=2,
    ):
        website = normalize_text(
            record.get("Sitio_internet")
        )

        if website:
            website_url = website

            if not website.lower().startswith(
                (
                    "http://",
                    "https://",
                )
            ):
                website_url = (
                    f"https://{website}"
                )

            website_cell = worksheet.cell(
                row=row_number,
                column=website_column,
            )

            website_cell.hyperlink = website_url
            website_cell.style = "Hyperlink"

        maps_url = normalize_text(
            record.get("Google_Maps_URL")
        )

        if maps_url:
            maps_cell = worksheet.cell(
                row=row_number,
                column=maps_column,
            )

            maps_cell.hyperlink = maps_url
            maps_cell.style = "Hyperlink"

    byte_stream = BytesIO()

    workbook.save(
        byte_stream
    )

    byte_stream.seek(0)

    return send_file(
        byte_stream,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        as_attachment=True,
        download_name=build_export_filename(
            "xlsx"
        ),
    )


def validate_form(
    form_data: dict[str, object],
) -> tuple[str, str, str, str, str, int]:
    entity = str(form_data["entity"]).strip()

    municipality = str(
        form_data["municipality"]
    ).strip()

    activity_class = str(
        form_data["activity_class"]
    ).strip()

    stratum = str(
        form_data["stratum"]
    ).strip()

    colony_filter = normalize_text(
        form_data["colony_filter"]
    )

    try:
        coverage_limit = int(
            str(form_data["maximum_records"])
        )
    except (TypeError, ValueError) as error:
        raise ValueError(
            "Selecciona una cobertura de búsqueda válida."
        ) from error

    valid_entities = get_valid_codes(
        ENTITY_OPTIONS
    )

    valid_municipalities = get_valid_codes(
        MUNICIPALITY_OPTIONS
    )

    valid_strata = get_valid_codes(
        STRATUM_OPTIONS
    )

    if entity not in valid_entities:
        raise ValueError(
            "Selecciona una entidad válida."
        )

    if municipality not in valid_municipalities:
        raise ValueError(
            "Selecciona una alcaldía válida."
        )

    if not (
        activity_class == "0"
        or (
            activity_class.isdigit()
            and len(activity_class) == 6
        )
    ):
        raise ValueError(
            "El código SCIAN debe contener seis dígitos "
            "o utilizar 0 para todas las actividades."
        )

    if stratum not in valid_strata:
        raise ValueError(
            "Selecciona un tamaño de negocio válido."
        )

    if coverage_limit not in VALID_COVERAGE_LIMITS:
        raise ValueError(
            "Selecciona una cobertura de búsqueda disponible."
        )

    if len(colony_filter) > 100:
        raise ValueError(
            "El filtro de colonia no puede superar "
            "los 100 caracteres."
        )

    return (
        entity,
        municipality,
        activity_class,
        stratum,
        colony_filter,
        coverage_limit,
    )


@app.route("/", methods=["GET", "POST"])
def index():
    form_data = get_default_form()

    results: list[dict[str, object]] = []

    error_message = ""
    api_result_count = 0
    search_executed = False
    search_description = ""

    denue_batch_count = 0
    coverage_limit_reached = False

    if request.method == "POST":
        search_executed = True

        action = request.form.get(
            "action",
            "search",
        ).strip()

        form_data = {
            "entity": request.form.get(
                "entity",
                "",
            ).strip(),
            "municipality": request.form.get(
                "municipality",
                "",
            ).strip(),
            "activity_class": request.form.get(
                "activity_class",
                "",
            ).strip(),
            "stratum": request.form.get(
                "stratum",
                "",
            ).strip(),
            "maximum_records": request.form.get(
                "maximum_records",
                "",
            ).strip(),
            "colony_filter": request.form.get(
                "colony_filter",
                "",
            ).strip(),
            "only_with_phone": (
                request.form.get("only_with_phone")
                == "on"
            ),
            "only_without_website": (
                request.form.get(
                    "only_without_website"
                )
                == "on"
            ),
        }

        try:
            if action not in {
                "search",
                "export_csv",
                "export_excel",
            }:
                raise ValueError(
                    "La acción solicitada no es válida."
                )

            (
                entity,
                municipality,
                activity_class,
                stratum,
                colony_filter,
                coverage_limit,
            ) = validate_form(form_data)

            (
                raw_results,
                denue_batch_count,
                coverage_limit_reached,
            ) = query_denue_coverage(
                entity=entity,
                municipality=municipality,
                activity_class=activity_class,
                stratum=stratum,
                coverage_limit=coverage_limit,
            )

            api_result_count = len(raw_results)

            results = filter_results(
                records=raw_results,
                only_with_phone=bool(
                    form_data["only_with_phone"]
                ),
                only_without_website=bool(
                    form_data["only_without_website"]
                ),
                colony_filter=colony_filter,
            )

            if action in {
                "export_csv",
                "export_excel",
            } and not results:
                raise ValueError(
                    "No hay prospectos para exportar "
                    "con los filtros seleccionados."
                )

            if action == "export_csv":
                return export_results_csv(
                    results
                )

            if action == "export_excel":
                return export_results_excel(
                    results
                )

            entity_name = get_option_name(
                ENTITY_OPTIONS,
                entity,
            )

            municipality_name = get_option_name(
                MUNICIPALITY_OPTIONS,
                municipality,
            )

            stratum_name = get_option_name(
                STRATUM_OPTIONS,
                stratum,
            )

            search_description = (
                f"{municipality_name}, {entity_name} · "
                f"SCIAN {activity_class} · "
                f"{stratum_name} · "
                f"Cobertura hasta {coverage_limit:,}"
            )

            if colony_filter:
                search_description += (
                    f" · Colonia contiene: {colony_filter}"
                )

        except ValueError as error:
            error_message = str(error)

        except requests.Timeout:
            error_message = (
                "La API del DENUE tardó demasiado "
                "en responder."
            )

        except requests.HTTPError as error:
            status_code = (
                error.response.status_code
                if error.response is not None
                else "desconocido"
            )

            error_message = (
                "La API del DENUE respondió con el "
                f"estado HTTP {status_code}."
            )

        except requests.RequestException:
            error_message = (
                "No fue posible conectarse con "
                "la API del DENUE."
            )

        except RuntimeError as error:
            error_message = str(error)

        except Exception:
            error_message = (
                "Ocurrió un error inesperado "
                "durante la consulta."
            )

    return render_template(
        "index.html",
        form_data=form_data,
        entity_options=ENTITY_OPTIONS,
        municipality_options=MUNICIPALITY_OPTIONS,
        stratum_options=STRATUM_OPTIONS,
        coverage_options=COVERAGE_OPTIONS,
        results=results,
        error_message=error_message,
        api_result_count=api_result_count,
        search_executed=search_executed,
        search_description=search_description,
        denue_batch_count=denue_batch_count,
        coverage_limit_reached=coverage_limit_reached,
    )


if __name__ == "__main__":
    app.run(
        host="127.0.0.1",
        port=5000,
        debug=True,
    )